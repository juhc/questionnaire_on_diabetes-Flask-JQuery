from flask import Blueprint, render_template, request, jsonify, send_file
from .models import Question, Answer, Group, Recomendations, Results, Texts
from . import db
import json
import openpyxl
import io
from config import Config
from collections import OrderedDict

home = Blueprint("home", __name__)


@home.route("/")
def index():
    tests = [group.name for group in Group.query.all()]
    return render_template("index.html", tests=tests)

@home.route("/test")
def test():
    answers = {"2":{"1":["1"],"2":["23"],"3":["awe"],"4":["2"],"5":["172"],"6":["72"],"7":["80"],"8":["3"],"9":["3"],"10":["3"]},"3":{"1":["1"],"2":["1"],"3":["2"],"4":["1"],"5":["2"]}}
    return get_attributes_fr(answers)

@home.route("/get-question", methods=["GET", "POST"])
def get_question_by_id():
    if request.method == "GET":
        response = {}
        group_id = request.args.get("group")

        if group_id:
            group = Group.query.get(group_id)

            if group.type == "question":
                response = get_question_data(group_id)
            if group.type == "text":
                response = get_text_data(group_id)
            if group.type == "recomendation":
                if group.name == "Рекомендации":
                    response = "recomendations"
                if group.name == "Оценка риска":
                    response = "risks"

        return jsonify(response)


@home.route("/get-recomendations", methods=["POST"])
def get_recomendations():
    answers = json.loads(request.data)
    response = recomendation_data(answers)
    save_result_to_db(answers)
    return jsonify(response)


@home.route("/get-risks", methods=["POST"])
def get_risks():
    answers = json.loads(request.data)
    response = risks_data(answers)
    return jsonify(response)


@home.route("/get-questions-count")
def get_questions_count():
    group_id = request.args.get("group")
    response = {"count": None}

    if group_id:
        count = len(Group.query.get(group_id).questions)
        response["count"] = count if count else None

    return jsonify(response)


@home.route("/get-tests-count")
def get_tests_count():
    response = {"count": len(Group.query.all())}

    return jsonify(response)


def get_question_data(group_id) -> dict:
    response = {}
    group = Group.query.get(group_id)
    questions = group.questions
    for i, q in enumerate(questions, start=1):
        question = dict.fromkeys(("text", "type", "notion", "answers"))
        question["text"] = (
            " ".join(("".join((str(i), ".")), q.text)) if group.id in (4,5,6) else q.text
        )
        question["notion"] = q.notion
        question["type"] = q.type

        answers = q.answers
        question["answers"] = [
            {"text": answer.text, "type": answer.type} for answer in answers
        ]

        response[len(response) + 1] = question
    response = {
        "group-type": "question",
        "data": response,
        "title": group.name,
    }
    return response


def recomendation_data(answers) -> dict:
    response = {}
    recomendations = get_questions_recomendations(answers)
    recomendations.append(get_debq_recomendations(answers))
    recomendations.append(get_dsmv_recomendations(answers))

    response = {
        key: {"title": "Рекомендовано", "text": value.text}
        for key, value in enumerate(recomendations, start=1)
    }
    response = {"group-type": "recomendation", "data": response}

    return response


def get_questions_and_answers_by_group(name):
    group = Group.query.filter(
        (Group.name == name) & (Group.type == "question")
    ).first()

    questions = group.questions

    return questions, group.id


def get_debq_recomendations(answers):
    questions, group_id = get_questions_and_answers_by_group(Config.DSM_V)
    answers = answers[str(group_id)]

    result = []
    for index, question in enumerate(questions, start=1):
        answer = int(answers[str(index)][0])
        result.append(
            Answer.query.filter_by(question_id=question.id).all()[answer - 1].point
        )

    if sum(result[0:10]) / 10 > 2.4:
        return Recomendations.query.filter_by(value="Ограничительный").first()

    elif sum(result[10:23]) / 13 > 1.8:
        return Recomendations.query.filter_by(value="Эмоциогенный").first()

    elif sum(result[23:33]) / 10 > 2.7:
        return Recomendations.query.filter_by(value="Экстернальный").first()

    else:
        return Recomendations.query.filter_by(value="Норма").first()


def get_dsmv_recomendations(answers):
    questions, group_id = get_questions_and_answers_by_group(Config.DSM_V)
    answers = answers[str(group_id)]

    yes_no_dict = {"yes": [], "no": []}

    for index, question in enumerate(questions, start=1):
        answer = answers[str(index)]
        answers_cur_question = question.answers
        for a in answer:
            a = int(a) - 1
            if "Да" in answers_cur_question[a].text:
                yes_no_dict["yes"].append(index)
                break
            elif "Нет" in answers_cur_question[a].text:
                yes_no_dict["no"].append(index)

    if all([yes_no_dict["yes"].count(i) > 0 for i in range(1, 7)]):
        return Recomendations.query.filter_by(value="Нервная булимия").first()

    elif (
        sum([1 for i in range(1, 14) if yes_no_dict["yes"].count(i) > 0])
        and 4 in yes_no_dict["no"]
    ):
        return Recomendations.query.filter_by(value="Компульсивное переедание").first()

    else:
        return Recomendations.query.filter_by(value="Неизвестное переедание").first()


def calculate_imt(answers, attributes) -> float:
    _, group_id = get_questions_and_answers_by_group(Config.PASSPORT)
    ans = (answers[str(group_id)])

    weight = int(attributes['weight'])
    height = int(attributes['height'])

    return weight / (height**2)


def get_persone_attributes(answers):
    questions, group_id = get_questions_and_answers_by_group(Config.PASSPORT)
    answers = answers[str(group_id)]
    attributes = {
        "Окружность талии на уровне пупка, см": "waist",
        "Вес, кг": "weight",
        "Рост, см": "height",
        "Возраст": "age",
        "Пол": "sex",
    }
    result = {}
    for i, v in enumerate(questions, start=1):
        if v.text in attributes:
            result[attributes[v.text]] = answers[str(i)][0]

    result['sex'] = Question.query.filter_by(text='Пол').first().answers[int(result['sex'][0])-1].text
    
    return result


def get_attributes_fr(answers):
    attributes = get_persone_attributes(answers)
    result = {}
    criteries = OrderedDict({45:0, 55:2, 65:3})
    for key in criteries:
        if int(attributes["age"][0]) - key < 0:
            result['age'] = criteries[key]
            break 
        result['age'] = 4
    
    criteries = OrderedDict({25:0, 30.001:1})
    imt = calculate_imt(answers, attributes)
    for key in criteries:
        if imt - key < 0:
            result['imt'] = criteries[key]
            break
        result['imt'] = 3
    
    criteries = OrderedDict({94:0, 102.001:1}) if attributes['sex'] == 'Мужской' else OrderedDict({80:0, 88.001:1})
    for key in criteries:
        if int(attributes['waist'][0]) - key < 0:
            result['waist'] = criteries[key]
            break
        result['waist'] = 3
    
    return result


def get_questions_recomendations(answers):
    questions = get_questions_and_answers_by_group(Config.DEBQ)[0]
    if answers["1"]["1"] == ["1"]:
        sex = "male"
    else:
        sex = "female"

    recomendations = []
    imt = calculate_imt(answers)
    waist = get_persone_attributes(answers)['waist']

    if sex == "male":
        if imt < 25 and waist < 94:
            recomendations.append(
                Recomendations.query.filter_by(value="Нормальный вес").first()
            )
        elif 25 <= imt <= 30 or 94 <= waist <= 102:
            recomendations.append(
                Recomendations.query.filter_by(value="Избыточный вес").first()
            )
        elif imt > 30 or waist > 102:
            recomendations.append(
                Recomendations.query.filter_by(value="Ожирение").first()
            )

    elif sex == "female":
        if imt < 25 and waist < 80:
            recomendations.append(
                Recomendations.query.filter_by(value="Нормальный вес").first()
            )
        elif 25 <= imt <= 30 or 80 <= waist <= 88:
            recomendations.append(
                Recomendations.query.filter_by(value="Избыточный вес").first()
            )
        elif imt > 30.0 or waist > 88:
            recomendations.append(
                Recomendations.query.filter_by(value="Ожирение").first()
            )
    answers = answers["2"]
    for index, question in enumerate(questions, start=1):
        value = int(answers[str(index)][0])

        answer = question.answers[value - 1]
        rec = Recomendations.query.filter(
            (Recomendations.extra == question.id)
            & (Recomendations.value == answer.text)
        ).first()

        if rec:
            recomendations.append(rec)

    return recomendations


def get_points_on_questions(answers) -> int:
    points = 0
    attributes_fr = get_attributes_fr(answers)
    group = Group.query.filter_by(name=Config.QUESTIONS).first()
    answers = answers[str(group.id)]
    questions = group.questions

    if len(questions) != len(answers):
        return 0

    for index, question in enumerate(questions, start=1):
        value = int(answers[str(index)][0])
        answer = question.answers[value - 1]
        points += answer.point

    points += sum(attributes_fr.values())

    return points


def risks_data(answers) -> dict:
    points = get_points_on_questions(answers)

    if points < 7:
        risk_level = "Низкий"
    elif 7 <= points <= 11:
        risk_level = "Слегка повышен"
    elif 12 <= points <= 14:
        risk_level = "Умеренный"
    elif 15 <= points <= 20:
        risk_level = "Высокий"
    else:
        risk_level = "Очень высокий"

    risk = Recomendations.query.filter_by(value=risk_level).first()

    response = {
        "group-type": "recomendation",
        "data": {
            "1": {
                "1": {
                    "title": "Уровень риска сахарного диабета 2 типа",
                    "text": risk.value,
                },
                "2": {"title": "Комментарий", "text": risk.text},
                "3": {
                    "title": "Вероятность развития сахарного диабета 2 типа в течение ближайших 10 лет",
                    "text": risk.extra,
                },
            }
        },
    }

    return response


def save_result_to_db(answers):
    result = Results(data=json.dumps(answers))
    db.session.add(result)
    db.session.commit()


@home.route("/results")
def get_results():
    book = save_results_to_excel()
    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        download_name="Результаты.xlsx",
        as_attachment=True,
    )


@home.route("/recomendations-xlsx")
def get_recomendations_xlsx():
    buffer = io.BytesIO()
    answers = json.loads(request.args.get("answers"))
    book = get_recomendations_file(answers)
    book.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="Рекомендации.xlsx")


def save_results_to_excel():
    results = Results.query.all()
    book = openpyxl.Workbook()
    sheet = book.active

    titles = get_questions_titles()

    for i, v in enumerate(titles, start=1):
        sheet.cell(row=1, column=i).value = v

    results = Results.query.all()

    for row, result in enumerate(results, start=2):
        data = json.loads(result.data)
        col = 1
        for group in data:
            for question in data[group]:
                questions_temp = group.questions[int(question) - 1]
                answers = questions_temp.answers
                if questions_temp.type == "radio":
                    if data[group][question][0] in [
                        str(i) for i in range(1, len(answers) + 1)
                    ]:
                        answer_title = answers[int(data[group][question][0]) - 1].text
                    else:
                        answer_title = data[group][question][0]
                    sheet.cell(row=row, column=col).value = answer_title
                elif questions_temp.type == "textbox":
                    sheet.cell(row=row, column=col).value = data[group][question][0]
                elif questions_temp.type == "checkbox":
                    answer_title = [
                        answers[int(i) - 1].text for i in data[group][question]
                    ]
                    sheet.cell(row=row, column=col).value = "\n".join(answer_title)

                col += 1

    return book


def get_questions_titles():
    questions = Question.query.order_by(Question.group_id).all()

    return [question.text for question in questions]


def get_recomendations_file(answers):
    recomendations = recomendation_data(answers)["data"]
    book = openpyxl.Workbook()
    sheet = book.active
    for i, v in enumerate(recomendations, start=1):
        sheet.cell(row=i, column=1).value = recomendations[v]["text"]

    risks = risks_data(answers)["data"]["1"]
    for row, key in enumerate(risks, start=len(recomendations) + 2):
        sheet.cell(row=row, column=1).value = risks[key]["title"]
        sheet.cell(row=row, column=2).value = risks[key]["text"]
    return book


def get_text_data(group_id):
    response = {
        "group-type": "text",
        "data": Texts.query.filter_by(group_id=group_id).first().text,
        "title": Group.query.get(group_id).name,
    }

    return response
